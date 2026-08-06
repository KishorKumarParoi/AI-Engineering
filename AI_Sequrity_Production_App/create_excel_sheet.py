import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def build_executive_course_catalog(output_file="Course_Catalog_Executive.xlsx"):
    """
    Generates a beautifully styled Excel workbook with an Executive Dark-Indigo Theme,
    custom KPI summary blocks, formatted dates, currency, and auto-adjusted column widths.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Courses Catalog"

    # Display gridlines for clean structure
    ws.views.sheetView[0].showGridLines = True

    # --- COLOR PALETTE (Modern Executive Slate & Indigo) ---
    COLOR_BANNER_BG  = "1E1B4B"  # Deep Indigo Header Banner
    COLOR_HEADER_BG  = "312E81"  # Table Header Fill
    COLOR_ZEBRA_ODD  = "F8FAFC"  # Ultra Light Slate
    COLOR_ZEBRA_EVEN = "FFFFFF"  # Pure White
    COLOR_KPI_BG     = "EEF2FF"  # Soft Indigo Tint for Summary Cards
    COLOR_TEXT_DARK  = "1E293B"  # Dark Slate Text
    COLOR_BORDER     = "CBD5E1"  # Light Gray Border
    COLOR_FREE_GREEN = "047857"  # Deep Emerald Green for Free Tag

    # --- FONTS ---
    font_banner_title = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
    font_banner_sub   = Font(name="Segoe UI", size=9, italic=True, color="C7D2FE")
    font_th           = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_tb           = Font(name="Segoe UI", size=10, color=COLOR_TEXT_DARK)
    font_free         = Font(name="Segoe UI", size=10, bold=True, color=COLOR_FREE_GREEN)
    font_kpi_lbl      = Font(name="Segoe UI", size=9, bold=True, color="4338CA")
    font_kpi_val      = Font(name="Segoe UI", size=13, bold=True, color="1E1B4B")

    # --- FILLS & BORDERS ---
    fill_banner = PatternFill(start_color=COLOR_BANNER_BG, end_color=COLOR_BANNER_BG, fill_type="solid")
    fill_header = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    fill_odd    = PatternFill(start_color=COLOR_ZEBRA_ODD, end_color=COLOR_ZEBRA_ODD, fill_type="solid")
    fill_even   = PatternFill(start_color=COLOR_ZEBRA_EVEN, end_color=COLOR_ZEBRA_EVEN, fill_type="solid")
    fill_kpi    = PatternFill(start_color=COLOR_KPI_BG, end_color=COLOR_KPI_BG, fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color=COLOR_BORDER),
        right=Side(style='thin', color=COLOR_BORDER),
        top=Side(style='thin', color=COLOR_BORDER),
        bottom=Side(style='thin', color=COLOR_BORDER)
    )

    # --- DATA SOURCE ---
    courses = [
        {
            "Course Name": "Mastering RAG & Multi-Agent Architecture",
            "Course Author": "Greg Loughnane",
            "Author Description": "Founder & AI Architect specializing in LLMOps & Production RAG systems",
            "Source": "AI Makerspace",
            "Price": 499.00,
            "Updated Date": "2026-07-15"
        },
        {
            "Course Name": "Advanced LLM Engineering & Quantization",
            "Course Author": "Dr. Sebastian Raschka",
            "Author Description": "Lead AI Researcher & Author of 'Build a Large Language Model (From Scratch)'",
            "Source": "Ahead of AI",
            "Price": 299.00,
            "Updated Date": "2026-06-20"
        },
        {
            "Course Name": "Production LangGraph & Multi-Turn Agents",
            "Course Author": "Harrison Chase",
            "Author Description": "CEO & Co-founder of LangChain",
            "Source": "DeepLearning.AI",
            "Price": 0.00,
            "Updated Date": "2026-08-01"
        },
        {
            "Course Name": "GCP Cloud AI Engineering & Vertex Scale",
            "Course Author": "Kishor Kumar Paroi",
            "Author Description": "Senior Cloud AI Engineer & System Architect",
            "Source": "Maven AI Engineering",
            "Price": 350.00,
            "Updated Date": "2026-08-04"
        },
        {
            "Course Name": "System Design for Generative AI Applications",
            "Course Author": "Chip Huyen",
            "Author Description": "Author of 'Designing Machine Learning Systems' & Founder of Claypot AI",
            "Source": "O'Reilly Media",
            "Price": 199.00,
            "Updated Date": "2026-05-10"
        },
        {
            "Course Name": "Full-Stack FastAPI & Next.js AI Integration",
            "Course Author": "Tiangolo (Sebastián Ramírez)",
            "Author Description": "Creator of FastAPI & SQLModel Frameworks",
            "Source": "FastAPI / GitHub",
            "Price": 0.00,
            "Updated Date": "2026-07-28"
        },
        {
            "Course Name": "Vector DBs & Hybrid Search with Qdrant",
            "Course Author": "David Andreis",
            "Author Description": "Developer Relations Lead & Vector DB Specialist at Qdrant",
            "Source": "Qdrant Academy",
            "Price": 0.00,
            "Updated Date": "2026-06-05"
        },
        {
            "Course Name": "Enterprise Security & Guardrails for LLMs",
            "Course Author": "LlamaIndex Team",
            "Author Description": "Core Maintainers of LlamaIndex & Guardrails AI",
            "Source": "LlamaIndex Learning",
            "Price": 150.00,
            "Updated Date": "2026-07-02"
        }
    ]

    # --- 1. TITLE BANNER (Rows 1 & 2) ---
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")

    ws["A1"] = "🎓 PROFESSIONAL COURSE CATALOG & KNOWLEDGE INDEX"
    ws["A1"].font = font_banner_title
    ws["A1"].fill = fill_banner
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A2"] = f"Last Updated: {datetime.now().strftime('%B %d, %Y')} | Curated High-Impact Learning Resources"
    ws["A2"].font = font_banner_sub
    ws["A2"].fill = fill_banner
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    # --- 2. SUMMARY KPI CARDS (Rows 4 & 5) ---
    kpi_configs = [
        ("A", "B", "TOTAL COURSES", str(len(courses))),
        ("C", "D", "AVG COURSE PRICE", f"${sum(c['Price'] for c in courses) / len(courses):.2f}"),
        ("E", "F", "FREE / PAID RATIO", f"{sum(1 for c in courses if c['Price'] == 0)} Free | {sum(1 for c in courses if c['Price'] > 0)} Paid")
    ]

    for start_col, end_col, title, val in kpi_configs:
        ws.merge_cells(f"{start_col}4:{end_col}4")
        ws.merge_cells(f"{start_col}5:{end_col}5")
        
        c_title = ws[f"{start_col}4"]
        c_val   = ws[f"{start_col}5"]
        
        c_title.value = title
        c_title.font  = font_kpi_lbl
        c_title.alignment = Alignment(horizontal="center", vertical="center")
        
        c_val.value = val
        c_val.font  = font_kpi_val
        c_val.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(4, 6):
        ws.row_dimensions[row].height = 20
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill_kpi
            cell.border = thin_border

    ws.row_dimensions[6].height = 10 # Spacing

    # --- 3. TABLE HEADERS (Row 7) ---
    headers = ["Course Name", "Course Author", "Author Description", "Source", "Price", "Updated Date"]
    ws.row_dimensions[7].height = 26

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_idx, value=header)
        cell.font = font_th
        cell.fill = fill_header
        cell.border = thin_border
        align_pos = "center" if header in ["Source", "Price", "Updated Date"] else "left"
        cell.alignment = Alignment(horizontal=align_pos, vertical="center")

    # --- 4. TABLE DATA (Row 8+) ---
    for row_idx, c in enumerate(courses, 8):
        ws.row_dimensions[row_idx].height = 23
        current_fill = fill_even if row_idx % 2 == 0 else fill_odd

        cell_name   = ws.cell(row=row_idx, column=1, value=c["Course Name"])
        cell_author = ws.cell(row=row_idx, column=2, value=c["Course Author"])
        cell_desc   = ws.cell(row=row_idx, column=3, value=c["Author Description"])
        cell_src    = ws.cell(row=row_idx, column=4, value=c["Source"])
        cell_price  = ws.cell(row=row_idx, column=5, value=c["Price"])
        cell_date   = ws.cell(row=row_idx, column=6, value=datetime.strptime(c["Updated Date"], "%Y-%m-%d").date())

        cell_name.alignment   = Alignment(horizontal="left", vertical="center")
        cell_author.alignment = Alignment(horizontal="left", vertical="center")
        cell_desc.alignment   = Alignment(horizontal="left", vertical="center")
        cell_src.alignment    = Alignment(horizontal="center", vertical="center")
        cell_date.alignment   = Alignment(horizontal="center", vertical="center")

        cell_date.number_format = "YYYY-MM-DD"

        if c["Price"] == 0:
            cell_price.value = "Free"
            cell_price.font = font_free
            cell_price.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell_price.number_format = '"$"#,##0.00'
            cell_price.font = font_tb
            cell_price.alignment = Alignment(horizontal="right", vertical="center")

        for cell in (cell_name, cell_author, cell_desc, cell_src, cell_date):
            cell.font = font_tb
            cell.fill = current_fill
            cell.border = thin_border

        if c["Price"] != 0:
            cell_price.fill = current_fill
            cell_price.border = thin_border
        else:
            cell_price.fill = current_fill
            cell_price.border = thin_border

    # --- 5. DYNAMIC COLUMN WIDTHS WITH PADDING ---
    min_widths = {"A": 36, "B": 26, "C": 58, "D": 22, "E": 15, "F": 16}
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col_letter].width = max(max_len + 3, min_widths.get(col_letter, 15))

    wb.save(output_file)
    print(f"✅ Excel sheet created successfully at: {output_file}")

if __name__ == "__main__":
    build_executive_course_catalog()
